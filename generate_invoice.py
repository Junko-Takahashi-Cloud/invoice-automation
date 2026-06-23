import pandas as pd
import openpyxl
import win32com.client as win32
from datetime import datetime
import os

INPUT_FILE = "input_data.xlsx"
TEMPLATE_FILE = "invoice_template.xlsx"
OUTPUT_DIR = "output_invoice"

os.makedirs(OUTPUT_DIR, exist_ok=True)

df = pd.read_excel(INPUT_FILE)

for index, row in df.iterrows():
    client = row["請求先"]
    work1 = row["品目1"]
    work2 = row["品目2"]
    work3 = row["品目3"]
    price = row["合計金額"]
    date = datetime.now().strftime("%Y-%m-%d")

    wb = openpyxl.load_workbook(TEMPLATE_FILE)
    ws = wb.active

    ws["B4"] = date
    ws["B6"] = client
    ws["A10"] = work1
    ws["A11"] = work2
    ws["A12"] = work3
    ws["C15"] = price

    excel_path = f"{OUTPUT_DIR}/請求書_{client}_{date}.xlsx"
    pdf_path = f"{OUTPUT_DIR}/請求書_{client}_{date}.pdf"

    wb.save(excel_path)

    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False
    wb_pdf = excel.Workbooks.Open(os.path.abspath(excel_path))
    wb_pdf.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
    wb_pdf.Close()
    excel.Quit()

print("請求書PDFの自動生成が完了しました。")
